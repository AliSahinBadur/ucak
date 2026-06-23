from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
import hashlib
from io import BytesIO, StringIO
from pathlib import Path
import re
import unicodedata
from urllib.parse import unquote, urlparse

from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from ..db.models import CatalogDocumentLink, Document, ReportCatalogEntry


@dataclass(frozen=True)
class CatalogRow:
    report_code: str
    vehicle_name: str
    report_title: str
    discipline: str
    report_date: str | None = None
    authors: str | None = None
    source_path: str | None = None


class CatalogService:
    EXPECTED_COLUMNS = 6

    def __init__(self, session: Session) -> None:
        self.session = session

    def import_bytes(self, file_name: str, content: bytes) -> dict:
        suffix = Path(file_name or "").suffix.lower()
        rows = self._parse_xlsx(content) if suffix == ".xlsx" else self._parse_text(content)
        created = 0
        duplicates = 0
        updated = 0
        errors: list[str] = []
        seen_hashes: set[str] = set()

        for index, row in enumerate(rows, start=1):
            try:
                row_hash = self._row_hash(row)
                if row_hash in seen_hashes:
                    duplicates += 1
                    continue
                seen_hashes.add(row_hash)

                existing = self.session.scalar(
                    select(ReportCatalogEntry).where(ReportCatalogEntry.row_hash == row_hash)
                )
                if existing:
                    if row.source_path and existing.source_path != row.source_path:
                        existing.source_path = row.source_path
                        updated += 1
                    duplicates += 1
                    continue

                self.session.add(
                    ReportCatalogEntry(
                        report_code=row.report_code,
                        vehicle_name=row.vehicle_name,
                        report_title=row.report_title,
                        discipline=row.discipline,
                        report_date=row.report_date,
                        authors=row.authors,
                        source_path=row.source_path,
                        row_hash=row_hash,
                    )
                )
                created += 1
            except ValueError as exc:
                errors.append(f"row {index}: {exc}")

        self.session.commit()
        return {
            "file_name": file_name,
            "rows_seen": len(rows),
            "created_count": created,
            "duplicate_count": duplicates,
            "updated_count": updated,
            "error_count": len(errors),
            "errors": errors[:20],
        }

    def search(self, query: str = "", vehicle: str = "", discipline: str = "", limit: int = 20) -> list[dict]:
        conditions = []
        if query.strip():
            query_like = f"%{query.strip()}%"
            conditions.append(
                or_(
                    ReportCatalogEntry.report_code.ilike(query_like),
                    ReportCatalogEntry.vehicle_name.ilike(query_like),
                    ReportCatalogEntry.report_title.ilike(query_like),
                    ReportCatalogEntry.discipline.ilike(query_like),
                    ReportCatalogEntry.authors.ilike(query_like),
                )
            )
        if vehicle.strip():
            conditions.append(ReportCatalogEntry.vehicle_name.ilike(f"%{vehicle.strip()}%"))
        if discipline.strip():
            conditions.append(ReportCatalogEntry.discipline.ilike(f"%{discipline.strip()}%"))

        statement = select(ReportCatalogEntry).order_by(ReportCatalogEntry.report_date.desc(), ReportCatalogEntry.id.desc())
        if conditions:
            statement = statement.where(*conditions)
        rows = self.session.execute(statement.limit(limit)).scalars().all()
        return [self._entry_payload(row) for row in rows]

    def answer_catalog_question(self, question: str, limit: int = 30) -> dict:
        profile = self._question_profile(question)
        if profile["intent"] == "analysis_type_summary":
            return self._answer_analysis_type_summary(question=question, profile=profile, limit=limit)
        if profile["intent"] == "vehicle_comparison":
            return self._answer_vehicle_comparison(question=question, profile=profile, limit=limit)
        if profile["intent"] == "vehicle_ranking":
            return self._answer_vehicle_ranking(question=question, profile=profile, limit=limit)

        candidates = self._catalog_candidates(profile, limit=limit)
        matched_documents = self._match_documents(candidates)
        answer = self._build_catalog_answer(question=question, profile=profile, candidates=candidates)
        return {
            "question": question,
            "answer": answer,
            "answer_found": bool(candidates),
            "match_count": len(candidates),
            "filters": {
                "vehicle": profile["vehicle"],
                "vehicles": profile["vehicles"],
                "discipline": profile["discipline"],
                "year": profile["year"],
                "intent": profile["intent"],
                "query_terms": profile["query_terms"],
            },
            "catalog_matches": [
                {
                    **self._entry_payload(entry),
                    "matched_document_id": matched_documents.get(entry.id),
                }
                for entry in candidates
            ],
        }

    def _catalog_candidates(self, profile: dict, limit: int) -> list[ReportCatalogEntry]:
        statement = select(ReportCatalogEntry)
        conditions = []
        if profile["vehicle"]:
            conditions.append(ReportCatalogEntry.vehicle_name.ilike(f"%{profile['vehicle']}%"))
        if profile["discipline"]:
            discipline = profile["discipline"]
            conditions.append(
                or_(
                    ReportCatalogEntry.discipline.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_title.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{discipline}%"),
                )
            )
        if profile["year"]:
            year = profile["year"]
            conditions.append(
                or_(
                    ReportCatalogEntry.report_date.ilike(f"{year}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{year}%"),
                )
            )
        for term in profile["query_terms"]:
            conditions.append(
                or_(
                    ReportCatalogEntry.report_code.ilike(f"%{term}%"),
                    ReportCatalogEntry.vehicle_name.ilike(f"%{term}%"),
                    ReportCatalogEntry.report_title.ilike(f"%{term}%"),
                    ReportCatalogEntry.discipline.ilike(f"%{term}%"),
                )
            )
        if conditions:
            statement = statement.where(*conditions)

        rows = self.session.execute(
            statement.order_by(ReportCatalogEntry.report_date.desc(), ReportCatalogEntry.id.desc()).limit(limit)
        ).scalars().all()
        return rows

    def _question_profile(self, question: str) -> dict:
        normalized = self._normalize_text(question)
        vehicles = self._detect_vehicles(normalized)
        vehicle = vehicles[0] if vehicles else ""
        discipline = self._detect_discipline(normalized)
        year_match = re.search(r"\b(20\d{2})\b", normalized)
        year = year_match.group(1) if year_match else ""
        normalized_vehicle = self._normalize_text(vehicle)
        compact_vehicle = normalized_vehicle.replace(" ", "")
        filler = {
            "araci",
            "arac",
            "ile",
            "kac",
            "tane",
            "adet",
            "test",
            "testi",
            "yapildi",
            "yapilmis",
            "rapor",
            "raporu",
            "raporlar",
            "raporlari",
            "var",
            "icin",
            "yil",
            "yili",
            "yilinda",
            "nelerdir",
            "hangi",
        }
        query_terms = [
            token
            for token in re.findall(r"\w+", normalized)
            if len(token) >= 3
            and token not in filler
            and token != year
            and token not in normalized_vehicle.split()
            and token not in compact_vehicle
            and token != discipline.casefold()
        ]
        return {
            "vehicle": vehicle,
            "vehicles": vehicles,
            "discipline": discipline,
            "year": year,
            "intent": self._detect_intent(normalized, vehicles=vehicles, discipline=discipline),
            "query_terms": query_terms[:4],
        }

    def _answer_analysis_type_summary(self, question: str, profile: dict, limit: int) -> dict:
        statement = (
            select(ReportCatalogEntry.discipline, func.count(ReportCatalogEntry.id))
            .where(ReportCatalogEntry.discipline != "")
        )
        conditions = []
        if profile["vehicle"]:
            conditions.append(ReportCatalogEntry.vehicle_name.ilike(f"%{profile['vehicle']}%"))
        if profile["year"]:
            year = profile["year"]
            conditions.append(
                or_(
                    ReportCatalogEntry.report_date.ilike(f"{year}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{year}%"),
                )
            )
        if conditions:
            statement = statement.where(*conditions)

        rows = self.session.execute(
            statement.group_by(ReportCatalogEntry.discipline).order_by(func.count(ReportCatalogEntry.id).desc())
        ).all()
        rows = [(discipline, count) for discipline, count in rows if discipline]

        if not rows:
            return {
                "question": question,
                "answer": "Katalogda bu soruya uygun analiz tipi bulunamadi.",
                "answer_found": False,
                "match_count": 0,
                "filters": {
                    "vehicle": profile["vehicle"],
                    "vehicles": profile["vehicles"],
                    "discipline": profile["discipline"],
                    "year": profile["year"],
                    "intent": profile["intent"],
                    "query_terms": profile["query_terms"],
                },
                "catalog_matches": [],
            }

        scope_parts = []
        if profile["vehicle"]:
            scope_parts.append(profile["vehicle"])
        if profile["year"]:
            scope_parts.append(profile["year"])
        scope = " / ".join(scope_parts) if scope_parts else "katalog"
        visible_rows = rows[:limit]
        lines = [f"{scope} icinde {len(rows)} farkli analiz tipi var:"]
        lines.extend(f"{index}. {discipline}: {count} rapor" for index, (discipline, count) in enumerate(visible_rows, start=1))
        return {
            "question": question,
            "answer": "\n".join(lines),
            "answer_found": True,
            "match_count": len(rows),
            "filters": {
                "vehicle": profile["vehicle"],
                "vehicles": profile["vehicles"],
                "discipline": profile["discipline"],
                "year": profile["year"],
                "intent": profile["intent"],
                "query_terms": profile["query_terms"],
            },
            "catalog_matches": [],
        }

    def _answer_vehicle_comparison(self, question: str, profile: dict, limit: int) -> dict:
        vehicles = profile["vehicles"]
        if len(vehicles) < 2:
            return self._answer_vehicle_ranking(question=question, profile=profile, limit=limit)

        lines = ["Kataloga gore arac karsilastirmasi:"]
        total_matches = 0
        for vehicle in vehicles[:6]:
            summary = self._vehicle_summary(vehicle=vehicle, discipline=profile["discipline"], year=profile["year"])
            total_matches += summary["total"]
            discipline_text = f" {profile['discipline']}" if profile["discipline"] else ""
            lines.append(f"- {vehicle}: {summary['total']}{discipline_text} raporu")
            if summary["breakdown"]:
                breakdown = ", ".join(f"{item['discipline']}={item['count']}" for item in summary["breakdown"][:5])
                lines.append(f"  Dagilim: {breakdown}")

        winner = self._comparison_winner(vehicles[:6], discipline=profile["discipline"], year=profile["year"])
        if winner:
            lines.append(f"En yuksek kayit: {winner['vehicle']} ({winner['count']} rapor).")

        return {
            "question": question,
            "answer": "\n".join(lines),
            "answer_found": total_matches > 0,
            "match_count": total_matches,
            "filters": self._filters_payload(profile),
            "catalog_matches": [],
        }

    def _answer_vehicle_ranking(self, question: str, profile: dict, limit: int) -> dict:
        statement = select(ReportCatalogEntry.vehicle_name, func.count(ReportCatalogEntry.id))
        conditions = []
        if profile["discipline"]:
            discipline = profile["discipline"]
            conditions.append(
                or_(
                    ReportCatalogEntry.discipline.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_title.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{discipline}%"),
                )
            )
        if profile["year"]:
            year = profile["year"]
            conditions.append(
                or_(
                    ReportCatalogEntry.report_date.ilike(f"{year}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{year}%"),
                )
            )
        if conditions:
            statement = statement.where(*conditions)

        rows = self.session.execute(
            statement.group_by(ReportCatalogEntry.vehicle_name)
            .order_by(func.count(ReportCatalogEntry.id).desc())
            .limit(limit)
        ).all()
        rows = [(vehicle, count) for vehicle, count in rows if vehicle]
        if not rows:
            return {
                "question": question,
                "answer": "Katalogda bu karsilastirma icin uygun arac kaydi bulunamadi.",
                "answer_found": False,
                "match_count": 0,
                "filters": self._filters_payload(profile),
                "catalog_matches": [],
            }

        scope = profile["discipline"] or "tum analiz tipleri"
        if profile["year"]:
            scope = f"{profile['year']} / {scope}"
        lines = [f"{scope} icin en cok raporu olan araclar:"]
        lines.extend(f"{index}. {vehicle}: {count} rapor" for index, (vehicle, count) in enumerate(rows, start=1))
        return {
            "question": question,
            "answer": "\n".join(lines),
            "answer_found": True,
            "match_count": len(rows),
            "filters": self._filters_payload(profile),
            "catalog_matches": [],
        }

    def _vehicle_summary(self, vehicle: str, discipline: str = "", year: str = "") -> dict:
        statement = select(ReportCatalogEntry.discipline, func.count(ReportCatalogEntry.id)).where(
            ReportCatalogEntry.vehicle_name.ilike(f"%{vehicle}%")
        )
        if discipline:
            statement = statement.where(
                or_(
                    ReportCatalogEntry.discipline.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_title.ilike(f"%{discipline}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{discipline}%"),
                )
            )
        if year:
            statement = statement.where(
                or_(
                    ReportCatalogEntry.report_date.ilike(f"{year}%"),
                    ReportCatalogEntry.report_code.ilike(f"%{year}%"),
                )
            )
        rows = self.session.execute(
            statement.group_by(ReportCatalogEntry.discipline).order_by(func.count(ReportCatalogEntry.id).desc())
        ).all()
        breakdown = [{"discipline": item_discipline, "count": count} for item_discipline, count in rows if item_discipline]
        return {"total": sum(item["count"] for item in breakdown), "breakdown": breakdown}

    def _comparison_winner(self, vehicles: list[str], discipline: str = "", year: str = "") -> dict | None:
        scores = [
            {"vehicle": vehicle, "count": self._vehicle_summary(vehicle=vehicle, discipline=discipline, year=year)["total"]}
            for vehicle in vehicles
        ]
        scores = [score for score in scores if score["count"] > 0]
        if not scores:
            return None
        return max(scores, key=lambda item: item["count"])

    def _detect_vehicles(self, normalized_question: str) -> list[str]:
        vehicle_rows = self.session.execute(select(ReportCatalogEntry.vehicle_name).distinct()).all()
        matches: list[tuple[int, str]] = []
        for (vehicle_name,) in vehicle_rows:
            normalized_vehicle = self._normalize_text(vehicle_name)
            variants = {
                normalized_vehicle,
                normalized_vehicle.replace(" ", ""),
                normalized_vehicle.replace("-", " "),
            }
            vehicle_tokens = [token for token in normalized_vehicle.split() if not token.startswith("gen")]
            if len(vehicle_tokens) >= 2:
                variants.add(" ".join(vehicle_tokens))
                variants.add("".join(vehicle_tokens))
            for variant in variants:
                if variant and variant in normalized_question:
                    matches.append((len(variant), vehicle_name))
                    break
        matches.sort(key=lambda item: item[0], reverse=True)
        vehicles: list[str] = []
        for _, vehicle_name in matches:
            normalized_candidate = self._normalize_text(vehicle_name)
            compact_candidate = normalized_candidate.replace(" ", "").replace("-", "").replace("_", "")
            if any(
                normalized_candidate in self._normalize_text(existing)
                or self._normalize_text(existing) in normalized_candidate
                or self._normalize_text(existing).replace(" ", "").replace("-", "").replace("_", "").startswith(compact_candidate)
                or compact_candidate.startswith(self._normalize_text(existing).replace(" ", "").replace("-", "").replace("_", ""))
                for existing in vehicles
            ):
                continue
            vehicles.append(vehicle_name)
        return vehicles[:8]

    def _detect_discipline(self, normalized_question: str) -> str:
        aliases = {
            "nvh": "NVH",
            "tase": "TASE",
            "durability": "DURABILITY",
            "dayanim": "DURABILITY",
            "dur": "DURABILITY",
            "safety": "SAFETY",
            "emniyet": "SAFETY",
            "cfd": "CFD",
            "akis": "CFD",
            "fatigue": "FATIGUE",
            "yorulma": "FATIGUE",
            "ved": "VED",
        }
        for token, discipline in aliases.items():
            if token in normalized_question:
                return discipline
        return ""

    @staticmethod
    def _detect_intent(normalized_question: str, vehicles: list[str], discipline: str = "") -> str:
        analysis_type_terms = (
            "analiz tipi",
            "analiz turu",
            "analiz turleri",
            "analysis type",
            "disiplin",
            "disiplinler",
            "kategori",
            "kategoriler",
        )
        if any(term in normalized_question for term in analysis_type_terms):
            return "analysis_type_summary"
        comparison_terms = ("karsilastir", "kiyasla", "fark", "farki", "daha fazla", "daha cok", "en cok", "en fazla")
        if len(vehicles) >= 2 and any(term in normalized_question for term in comparison_terms):
            return "vehicle_comparison"
        if any(term in normalized_question for term in ("hangi arac", "hangi araclar", "en cok", "en fazla", "daha fazla")):
            return "vehicle_ranking"
        return "catalog_search"

    @staticmethod
    def _filters_payload(profile: dict) -> dict:
        return {
            "vehicle": profile["vehicle"],
            "vehicles": profile["vehicles"],
            "discipline": profile["discipline"],
            "year": profile["year"],
            "intent": profile["intent"],
            "query_terms": profile["query_terms"],
        }

    def _build_catalog_answer(self, question: str, profile: dict, candidates: list[ReportCatalogEntry]) -> str:
        if not candidates:
            return "Katalogda bu soruya uygun rapor kaydi bulunamadi."

        normalized = self._normalize_text(question)
        wants_count = any(token in normalized for token in ("kac", "adet", "sayisi", "tane"))
        vehicle_text = profile["vehicle"] or "ilgili arac"
        discipline_text = profile["discipline"] or "ilgili konu"
        if wants_count:
            return f"{vehicle_text} icin katalogda {discipline_text} kapsaminda {len(candidates)} rapor kaydi bulundu."

        lines = [f"{len(candidates)} katalog kaydi bulundu:"]
        for index, entry in enumerate(candidates[:8], start=1):
            date_text = f" | {entry.report_date}" if entry.report_date else ""
            lines.append(f"{index}. {entry.report_code} - {entry.vehicle_name} - {entry.report_title} ({entry.discipline}){date_text}")
        return "\n".join(lines)

    def _match_documents(self, entries: list[ReportCatalogEntry]) -> dict[int, int]:
        if not entries:
            return {}
        entry_ids = [entry.id for entry in entries]
        explicit_links = self.session.execute(
            select(CatalogDocumentLink.catalog_entry_id, CatalogDocumentLink.document_id)
            .where(CatalogDocumentLink.catalog_entry_id.in_(entry_ids))
        ).all()
        matches: dict[int, int] = {
            int(catalog_entry_id): int(document_id)
            for catalog_entry_id, document_id in explicit_links
        }
        documents = self.session.execute(select(Document.id, Document.title, Document.file_name)).all()
        for entry in entries:
            if entry.id in matches:
                continue
            entry_keys = self._document_match_keys(entry)
            for document in documents:
                document_text = self._normalize_text(f"{document.title} {document.file_name}")
                compact_document_text = self._compact_key(document_text)
                if any(key and key in document_text for key in entry_keys):
                    matches[entry.id] = int(document.id)
                    break
                if any(key and key in compact_document_text for key in entry_keys):
                    matches[entry.id] = int(document.id)
                    break
        return matches

    def _document_match_keys(self, entry: ReportCatalogEntry) -> set[str]:
        raw_values = [
            entry.report_code or "",
            entry.source_path or "",
        ]
        normalized_title = self._normalize_text(entry.report_title or "")
        keys = {
            normalized_title,
            self._compact_key(normalized_title),
        }

        for raw_value in raw_values:
            for variant in self._text_variants(raw_value):
                if not variant:
                    continue
                path_name = Path(variant).name
                normalized_value = self._normalize_text(variant)
                normalized_path_name = self._normalize_text(path_name)
                keys.update(
                    {
                        normalized_value,
                        normalized_path_name,
                        self._compact_key(normalized_value),
                        self._compact_key(normalized_path_name),
                    }
                )

                code_parts = [
                    part for part in re.split(r"[^a-z0-9]+", normalized_path_name)
                    if len(part) >= 2 and part not in {"rev", "pdf", "docx"}
                ]
                if len(code_parts) >= 3:
                    keys.add(self._compact_key(" ".join(code_parts[:4])))
                    keys.add(self._compact_key(" ".join(code_parts)))

        return {key for key in keys if len(key) >= 5}

    @staticmethod
    def _text_variants(value: str) -> list[str]:
        variants = [value]
        for encoding in ("latin1", "cp1252"):
            try:
                repaired = value.encode(encoding).decode("utf-8")
            except UnicodeError:
                continue
            if repaired and repaired not in variants:
                variants.append(repaired)
        return variants

    @staticmethod
    def _compact_key(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", text)

    def _parse_xlsx(self, content: bytes) -> list[CatalogRow]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX okumak icin openpyxl kurulu olmali. CSV/TSV de yukleyebilirsin.") from exc

        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        parsed: list[CatalogRow] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                values = [cell.value for cell in row]
                hyperlinks = [
                    cell.hyperlink.target
                    if cell.hyperlink and cell.hyperlink.target
                    else ""
                    for cell in row
                ]
                parsed_row = self._row_from_values(values, hyperlinks=hyperlinks)
                if parsed_row:
                    parsed.append(parsed_row)
        return parsed

    def _parse_text(self, content: bytes) -> list[CatalogRow]:
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        delimiter = "\t" if "\t" in sample else ";"
        if "," in sample and sample.count(",") > sample.count(";"):
            delimiter = ","
        reader = csv.reader(StringIO(text), delimiter=delimiter)
        parsed: list[CatalogRow] = []
        for row in reader:
            parsed_row = self._row_from_values(row)
            if parsed_row:
                parsed.append(parsed_row)
        return parsed

    def _row_from_values(self, values: tuple | list | None, hyperlinks: list[str] | None = None) -> CatalogRow | None:
        if not values:
            return None
        cells = [self._cell_to_text(value) for value in values]
        if not any(cells):
            return None
        if self._looks_like_header(cells):
            return None
        if len(cells) < self.EXPECTED_COLUMNS:
            return None
        if not all(cells[index] for index in range(4)):
            return None

        return CatalogRow(
            report_code=cells[0],
            vehicle_name=cells[1],
            report_title=cells[2],
            discipline=cells[3].upper(),
            report_date=cells[4] or None,
            authors=cells[5] or None,
            source_path=self._source_path_from_row(cells, hyperlinks or []),
        )

    @classmethod
    def _source_path_from_row(cls, cells: list[str], hyperlinks: list[str]) -> str | None:
        candidates = [cells[0] if cells else ""]
        candidates.extend(hyperlinks)
        for candidate in candidates:
            path = cls._normalize_source_path(candidate)
            if path:
                return path
        return None

    @staticmethod
    def _normalize_source_path(value: str) -> str | None:
        raw_value = str(value or "").strip().strip('"')
        if not raw_value or raw_value.startswith("#"):
            return None

        if raw_value.casefold().startswith("file:"):
            parsed = urlparse(raw_value)
            raw_value = unquote(parsed.path or parsed.netloc)
            if re.match(r"^/[A-Za-z]:/", raw_value):
                raw_value = raw_value[1:]
            raw_value = raw_value.replace("/", "\\")

        if "\\" not in raw_value and "/" not in raw_value:
            return None
        return raw_value

    @staticmethod
    def _cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime | date):
            return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
        return str(value).strip()

    @staticmethod
    def _looks_like_header(cells: list[str]) -> bool:
        normalized = CatalogService._normalize_text(" ".join(cells[:6]))
        return any(
            token in normalized
            for token in (
                "report name",
                "vehicle",
                "design criteria",
                "analysis type",
                "rapor kod",
                "arac",
                "baslik",
                "disiplin",
                "hazirlayan",
            )
        )

    @staticmethod
    def _row_hash(row: CatalogRow) -> str:
        payload = "|".join(
            [
                row.report_code,
                row.vehicle_name,
                row.report_title,
                row.discipline,
                row.report_date or "",
                row.authors or "",
            ]
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @staticmethod
    def _entry_payload(entry: ReportCatalogEntry) -> dict:
        return {
            "id": entry.id,
            "report_code": entry.report_code,
            "vehicle_name": entry.vehicle_name,
            "report_title": entry.report_title,
            "discipline": entry.discipline,
            "report_date": entry.report_date,
            "authors": entry.authors,
            "source_path": entry.source_path,
        }

    @staticmethod
    def _normalize_text(text: str) -> str:
        lowered = text.casefold().translate(
            str.maketrans(
                {
                    "ı": "i",
                    "ğ": "g",
                    "ü": "u",
                    "ş": "s",
                    "ö": "o",
                    "ç": "c",
                }
            )
        )
        normalized = unicodedata.normalize("NFKD", lowered)
        return "".join(char for char in normalized if not unicodedata.combining(char))
